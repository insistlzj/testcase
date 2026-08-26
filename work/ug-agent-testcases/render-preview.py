from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    candidates = [
        Path(r"C:\Windows\Fonts\msyhbd.ttc" if bold else r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simhei.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size)
    return ImageFont.load_default()


def argb_to_rgb(value: str | None, fallback: str) -> str:
    if not value:
        return fallback
    value = value[-6:]
    return f"#{value}"


def wrap_text(draw: ImageDraw.ImageDraw, text: str, text_font: ImageFont.ImageFont, max_width: int) -> list[str]:
    output: list[str] = []
    for paragraph in str(text or "").splitlines() or [""]:
        current = ""
        for char in paragraph:
            candidate = current + char
            if current and draw.textbbox((0, 0), candidate, font=text_font)[2] > max_width:
                output.append(current)
                current = char
            else:
                current = candidate
        output.append(current)
    return output or [""]


def main() -> None:
    workbook_path = Path(sys.argv[1])
    image_path = Path(sys.argv[2])
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    ws = wb["功能测试用例"]

    max_row = min(ws.max_row, 9)
    scale = 5.2
    column_widths = [
        max(70, int((ws.column_dimensions[chr(65 + index)].width or 12) * scale))
        for index in range(10)
    ]
    row_heights = [
        max(40, int((ws.row_dimensions[row].height or 24) * 1.12))
        for row in range(1, max_row + 1)
    ]
    margin = 18
    width = sum(column_widths) + margin * 2
    height = sum(row_heights) + margin * 2
    image = Image.new("RGB", (width, height), "#F4F7FA")
    draw = ImageDraw.Draw(image)
    header_font = font(16, bold=True)
    body_font = font(13)

    y = margin
    for row in range(1, max_row + 1):
        x = margin
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell_width = column_widths[col - 1]
            cell_height = row_heights[row - 1]
            if row == 1:
                fill = "#1F4E78"
                text_color = "#FFFFFF"
                selected_font = header_font
            else:
                fill = "#FFFFFF" if row % 2 == 0 else "#F7FAFC"
                text_color = "#1F2937"
                selected_font = body_font
            draw.rectangle(
                [x, y, x + cell_width, y + cell_height],
                fill=fill,
                outline="#B8C7D9",
                width=1,
            )
            lines = wrap_text(draw, str(cell.value or ""), selected_font, cell_width - 10)
            line_height = draw.textbbox((0, 0), "测Ag", font=selected_font)[3] + 3
            max_lines = max(1, (cell_height - 8) // line_height)
            visible = lines[:max_lines]
            if len(lines) > max_lines and visible:
                visible[-1] = (visible[-1][:-1] + "…") if len(visible[-1]) > 1 else "…"
            block_height = len(visible) * line_height
            text_y = y + max(4, (cell_height - block_height) // 2)
            for line in visible:
                bbox = draw.textbbox((0, 0), line, font=selected_font)
                text_width = bbox[2] - bbox[0]
                draw.text(
                    (x + max(5, (cell_width - text_width) // 2), text_y),
                    line,
                    font=selected_font,
                    fill=text_color,
                )
                text_y += line_height
            x += cell_width
        y += row_heights[row - 1]

    image_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(image_path)
    print(f"preview={image_path}")
    print(f"size={width}x{height}")


if __name__ == "__main__":
    main()
